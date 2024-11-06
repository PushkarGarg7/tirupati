# import os
# from dotenv import load_dotenv
# import json
# import google.generativeai as genai
# import openpyxl
# from openpyxl import load_workbook
# import random

# google_api_key = os.getenv('GOOGLE_API_KEY')
# load_dotenv()
# genai.configure(api_key=google_api_key)
# model = genai.GenerativeModel("gemini-1.5-flash")




# sample_file = genai.upload_file(path="invoice.pdf", display_name="test")
# prompt = """
# Please provide the invoice details in JSON format. The JSON object should contain:

# invoice_number: the invoice number as a string.
# invoice_date: the date of the invoice as a string.
# Items: a dictionary where each key is an item's code and the corresponding value is the item's name.

# just give these things that are asked nothing else, not even ``` json etc
# """

# response = model.generate_content([sample_file, prompt])


# result = response.text
# data = json.loads(result)

# invoice_number = data.get('invoice_number', 'Not Available')
# invoice_date = data.get('invoice_date', 'Not Available')
# items = data.get('Items', {})  

# print("Invoice Number:", invoice_number)
# print("Invoice Date:", invoice_date)
# print("Items Dictionary:", items)
# genai.delete_file(sample_file.name)


# templates_folder = "./"  

# output_folder = "../outputs"

# if not os.path.exists(output_folder):
#     os.makedirs(output_folder)

# def set_cell_value(sheet, cell, value):
#     for merged_range in sheet.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             top_left_cell = sheet[merged_range.bounds[0], merged_range.bounds[1]]
#             top_left_cell.value = value
#             return
    
#     cell.value = value

# def apply_random_adjustment(value):
#     if isinstance(value, (int, float)):  
#         adjustment = random.uniform(-0.1, 0.1)  
#         return value + adjustment
#     return value  
# for code, name in items.items():
    
#     template_filename = f"{code}.xlsx"
#     template_path = os.path.join(templates_folder, template_filename)
    
#     if os.path.exists(template_path):
#         workbook = load_workbook(template_path)
#         sheet = workbook.active  
        
#         set_cell_value(sheet, sheet["B4"], invoice_number)  
#         set_cell_value(sheet, sheet["C4"], invoice_date)    
#         row = 11  
#         while sheet[f"B{row}"].value:  
#             b_value = sheet[f"B{row}"].value  

#             d_value = apply_random_adjustment(b_value) if isinstance(b_value, (int, float)) else b_value
#             set_cell_value(sheet, sheet[f"D{row}"], d_value)  

#             row += 1 

#         output_filename = f"{code}_invoice.xlsx"
#         output_path = os.path.join(output_folder, output_filename)
#         workbook.save(output_path)
#         print(f"Excel file saved for item {name} with code {code} at {output_path}")
#     else:
#         print(f"Template for {code} not found.")
import os
from dotenv import load_dotenv
import json
import google.generativeai as genai
import openpyxl
from openpyxl import load_workbook
import random
import shutil

# Load environment variables
load_dotenv()
google_api_key = os.getenv('GOOGLE_API_KEY')

# Configure the Google API
genai.configure(api_key=google_api_key)
model = genai.GenerativeModel("gemini-1.5-flash")

# Function to process the invoice and return the paths of the saved XLS files
def process_invoice(invoice_path):
    try:
        sample_file = genai.upload_file(path=invoice_path, display_name="test")
        
        prompt = """
        Please provide the invoice details in JSON format. The JSON object should contain:
        invoice_number: the invoice number as a string.
        invoice_date: the date of the invoice as a string.
        Items: a dictionary where each key is an item's code and the corresponding value is the item's name.
        just give these things that are asked nothing else, not even ``` json etc
        """

        # Send the request to the model
        response = model.generate_content([sample_file, prompt])

        result = response.text
        data = json.loads(result)

        invoice_number = data.get('invoice_number', 'Not Available')
        invoice_date = data.get('invoice_date', 'Not Available')
        items = data.get('Items', {})  

        # Cleanup: delete the uploaded file from the model server
        genai.delete_file(sample_file.name)

        # Process each item
        output_folder = "./outputs"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        def set_cell_value(sheet, cell, value):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = sheet[merged_range.bounds[0], merged_range.bounds[1]]
                    top_left_cell.value = value
                    return
            cell.value = value

        def apply_random_adjustment(value):
            if isinstance(value, (int, float)):
                adjustment = random.uniform(-0.1, 0.1)
                return value + adjustment
            return value

        output_files = []

        for code, name in items.items():
            template_filename = f"{code}.xlsx"
            # template_path = os.path.join("templates", template_filename)
            # template_directory = os.path.join(os.getcwd(), "backend")  # Adjust the path if needed
            template_directory = os.path.join(os.path.dirname(os.getcwd()), "backend")

            # template_path = os.path.join(template_directory, template_directory)
            template_path = os.path.join(template_directory, template_filename)

            print(template_path)


            if os.path.exists(template_path):
                workbook = load_workbook(template_path)
                sheet = workbook.active  
                
                set_cell_value(sheet, sheet["B4"], invoice_number)  
                set_cell_value(sheet, sheet["C4"], invoice_date)    
                row = 11  
                while sheet[f"B{row}"].value:  
                    b_value = sheet[f"B{row}"].value  
                    d_value = apply_random_adjustment(b_value) if isinstance(b_value, (int, float)) else b_value
                    set_cell_value(sheet, sheet[f"D{row}"], d_value)  
                    row += 1 

                output_filename = f"{code}_invoice.xlsx"
                output_path = os.path.join(output_folder, output_filename)
                workbook.save(output_path)
                output_files.append(output_path)
            else:
                print(f"Template for {code} not found.")
        
        return output_files
    except Exception as e:
        print(f"Error processing invoice: {e}")
        return None
